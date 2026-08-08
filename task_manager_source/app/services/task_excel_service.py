from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models.task_entity import TaskEntity


class TaskExcelService:
    HEADERS = [
        "Task ID",
        "Task Number",
        "Parent Task ID",
        "Short Description",
        "Description",
        "Status",
        "Priority",
        "Task Type",
        "Category",
        "Tags",
        "Planned Start Date",
        "Planned End Date",
        "Actual Start Date",
        "Actual End Date",
        "Due Date",
        "Estimated Effort Hours",
        "Actual Effort Hours",
        "Remaining Effort Hours",
        "Progress Percentage",
        "Assigned To",
        "Team",
        "Blocked Reason",
        "External Reference",
        "Created By",
        "Created At",
        "Updated At",
        "Is Active",
    ]

    def __init__(self, database: Session) -> None:
        self.database = database

    def export_tasks(self) -> BytesIO:
        statement = (
            select(TaskEntity)
            .order_by(TaskEntity.task_id)
        )

        tasks = list(
            self.database.scalars(statement).all()
        )

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Tasks"

        worksheet.append(self.HEADERS)

        for task in tasks:
            worksheet.append(
                [
                    task.task_id,
                    task.task_number,
                    task.parent_task_id,
                    task.short_description,
                    task.description,
                    self._enum_value(task.status),
                    self._enum_value(task.priority),
                    self._enum_value(task.task_type),
                    task.category,
                    self._format_tags(task.tags),
                    task.planned_start_date,
                    task.planned_end_date,
                    task.actual_start_date,
                    task.actual_end_date,
                    task.due_date,
                    task.estimated_effort_hours,
                    task.actual_effort_hours,
                    task.remaining_effort_hours,
                    task.progress_percentage,
                    task.assigned_to,
                    task.team,
                    task.blocked_reason,
                    task.external_reference,
                    task.created_by,
                    task.created_at,
                    task.updated_at,
                    task.is_active,
                ]
            )

        self._format_worksheet(worksheet)

        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        return output

    @staticmethod
    def _enum_value(value: object) -> object:
        return getattr(value, "value", value)

    @staticmethod
    def _format_tags(tags: object) -> str:
        if tags is None:
            return ""

        if isinstance(tags, list):
            return ", ".join(str(tag) for tag in tags)

        return str(tags)

    @staticmethod
    def _format_worksheet(worksheet) -> None:
        header_fill = PatternFill(
            fill_type="solid",
            fgColor="1F4E78",
        )

        header_font = Font(
            color="FFFFFF",
            bold=True,
        )

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
            )

        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions

        date_columns = {
            11,
            12,
            13,
            14,
            15,
            25,
            26,
        }

        for row in worksheet.iter_rows(
            min_row=2,
            max_row=worksheet.max_row,
        ):
            for column_number in date_columns:
                row[column_number - 1].number_format = (
                    "dd-mmm-yyyy hh:mm"
                )

        preferred_widths = {
            1: 10,
            2: 14,
            3: 16,
            4: 35,
            5: 50,
            6: 16,
            7: 12,
            8: 14,
            9: 18,
            10: 25,
        }

        for column_number in range(
            1,
            worksheet.max_column + 1,
        ):
            width = preferred_widths.get(
                column_number,
                20,
            )

            worksheet.column_dimensions[
                get_column_letter(column_number)
            ].width = width

        worksheet.row_dimensions[1].height = 24