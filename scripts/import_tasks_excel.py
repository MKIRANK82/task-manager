from __future__ import annotations

import argparse
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from app.config.database import SessionLocal
from app.models.task import (
    TaskCreate,
    TaskPriority,
    TaskStatus,
    TaskType,
)
from app.models.task_activity import (
    ActivityType,
    TaskActivityCreate,
)
from app.repositories.task_activity_repository import (
    TaskActivityRepository,
)
from app.repositories.task_repository import (
    TaskRepository,
)
from app.services.task_activity_service import (
    TaskActivityService,
)
from app.services.task_service import TaskService


@dataclass
class ImportRow:
    excel_row: int
    upload_ref: int
    parent_upload_ref: int
    existing_parent_task_id: int

    short_description: str
    description: str | None

    status: TaskStatus
    priority: TaskPriority
    task_type: TaskType

    category: str | None
    tags: list[str]

    planned_start_date: datetime | None
    planned_end_date: datetime | None
    due_date: datetime | None

    estimated_effort_hours: float | None
    assigned_to: str | None
    team: str | None
    created_by: str


REQUIRED_HEADERS = [
    "Upload Ref",
    "Parent Upload Ref",
    "Existing Parent Task ID",
    "Short Description",
    "Description",
    "Status",
    "Priority",
    "Task Type",
    "Category",
    "Tags",
    "Planned Start Date",
    "Planned End Date",
    "Due Date",
    "Estimated Effort Hours",
    "Assigned To",
    "Team",
    "Created By",
]


def clean_text(value: Any) -> str | None:
    if value is None:
        return None

    cleaned = str(value).strip()
    return cleaned or None


def parse_integer(
    value: Any,
    *,
    default: int = 0,
) -> int:
    if value in (None, ""):
        return default

    if isinstance(value, bool):
        raise ValueError(
            f"Invalid integer value: {value}"
        )

    return int(value)


def parse_float(value: Any) -> float | None:
    if value in (None, ""):
        return None

    return float(value)


def parse_datetime(
    value: Any,
) -> datetime | None:
    if value in (None, ""):
        return None

    if isinstance(value, datetime):
        return value

    if isinstance(value, date):
        return datetime.combine(
            value,
            datetime.min.time(),
        )

    text = str(value).strip()

    accepted_formats = [
        "%Y-%m-%d",
        "%Y-%m-%d %H:%M",
        "%d-%m-%Y",
        "%d-%m-%Y %H:%M",
        "%d/%m/%Y",
        "%d/%m/%Y %H:%M",
    ]

    for date_format in accepted_formats:
        try:
            return datetime.strptime(
                text,
                date_format,
            )
        except ValueError:
            continue

    raise ValueError(
        f"Invalid date value: {value}"
    )


def parse_tags(value: Any) -> list[str]:
    text = clean_text(value)

    if text is None:
        return []

    return [
        tag.strip()
        for tag in text.split(",")
        if tag.strip()
    ]


def read_import_rows(
    file_path: Path,
) -> list[ImportRow]:
    workbook = load_workbook(
        filename=file_path,
        data_only=True,
    )

    if "Tasks" not in workbook.sheetnames:
        raise ValueError(
            "Workbook must contain a sheet named 'Tasks'."
        )

    worksheet = workbook["Tasks"]

    actual_headers = [
        cell.value
        for cell in worksheet[1]
    ]

    missing_headers = [
        header
        for header in REQUIRED_HEADERS
        if header not in actual_headers
    ]

    if missing_headers:
        raise ValueError(
            "Missing required columns: "
            + ", ".join(missing_headers)
        )

    header_indexes = {
        header: actual_headers.index(header) + 1
        for header in REQUIRED_HEADERS
    }

    rows: list[ImportRow] = []

    for row_number in range(
        2,
        worksheet.max_row + 1,
    ):
        def value(column_name: str) -> Any:
            return worksheet.cell(
                row=row_number,
                column=header_indexes[column_name],
            ).value

        short_description = clean_text(
            value("Short Description")
        )

        if short_description is None:
            empty_row = all(
                worksheet.cell(
                    row=row_number,
                    column=column_number,
                ).value in (None, "")
                for column_number in range(
                    1,
                    worksheet.max_column + 1,
                )
            )

            if empty_row:
                continue

            raise ValueError(
                f"Row {row_number}: "
                "Short Description is required."
            )

        upload_ref = parse_integer(
            value("Upload Ref")
        )

        if upload_ref <= 0:
            raise ValueError(
                f"Row {row_number}: "
                "Upload Ref must be greater than zero."
            )

        parent_upload_ref = parse_integer(
            value("Parent Upload Ref")
        )

        existing_parent_task_id = parse_integer(
            value("Existing Parent Task ID")
        )

        if (
            parent_upload_ref > 0
            and existing_parent_task_id > 0
        ):
            raise ValueError(
                f"Row {row_number}: specify either "
                "Parent Upload Ref or Existing Parent "
                "Task ID, not both."
            )

        status_text = (
            clean_text(value("Status"))
            or TaskStatus.NOT_STARTED.value
        )

        priority_text = (
            clean_text(value("Priority"))
            or TaskPriority.MEDIUM.value
        )

        task_type_text = (
            clean_text(value("Task Type"))
            or TaskType.TASK.value
        )

        planned_start = parse_datetime(
            value("Planned Start Date")
        )

        planned_end = parse_datetime(
            value("Planned End Date")
        )

        if (
            planned_start is not None
            and planned_end is not None
            and planned_end < planned_start
        ):
            raise ValueError(
                f"Row {row_number}: Planned End Date "
                "cannot be earlier than Planned Start Date."
            )

        rows.append(
            ImportRow(
                excel_row=row_number,
                upload_ref=upload_ref,
                parent_upload_ref=parent_upload_ref,
                existing_parent_task_id=(
                    existing_parent_task_id
                ),
                short_description=short_description,
                description=clean_text(
                    value("Description")
                ),
                status=TaskStatus(status_text.lower()),
                priority=TaskPriority(
                    priority_text.lower()
                ),
                task_type=TaskType(
                    task_type_text.lower()
                ),
                category=clean_text(
                    value("Category")
                ),
                tags=parse_tags(value("Tags")),
                planned_start_date=planned_start,
                planned_end_date=planned_end,
                due_date=parse_datetime(
                    value("Due Date")
                ),
                estimated_effort_hours=parse_float(
                    value("Estimated Effort Hours")
                ),
                assigned_to=clean_text(
                    value("Assigned To")
                ),
                team=clean_text(value("Team")),
                created_by=(
                    clean_text(value("Created By"))
                    or "Kiran"
                ),
            )
        )

    validate_upload_references(rows)

    return rows


def validate_upload_references(
    rows: list[ImportRow],
) -> None:
    upload_refs = [
        row.upload_ref
        for row in rows
    ]

    duplicate_refs = {
        upload_ref
        for upload_ref in upload_refs
        if upload_refs.count(upload_ref) > 1
    }

    if duplicate_refs:
        raise ValueError(
            "Duplicate Upload Ref values: "
            + ", ".join(
                str(value)
                for value in sorted(duplicate_refs)
            )
        )

    known_refs = set(upload_refs)

    for row in rows:
        if (
            row.parent_upload_ref > 0
            and row.parent_upload_ref
            not in known_refs
        ):
            raise ValueError(
                f"Row {row.excel_row}: Parent Upload Ref "
                f"{row.parent_upload_ref} does not exist."
            )

        if row.parent_upload_ref == row.upload_ref:
            raise ValueError(
                f"Row {row.excel_row}: a task cannot "
                "be its own parent."
            )


def create_services(database):
    task_repository = TaskRepository(database)

    activity_repository = (
        TaskActivityRepository(database)
    )

    activity_service = TaskActivityService(
        activity_repository
    )

    task_service = TaskService(
        repository=task_repository,
        activity_service=activity_service,
    )

    return task_service, activity_service


def validate_existing_parents(
    rows: list[ImportRow],
    task_service: TaskService,
) -> None:
    for row in rows:
        parent_id = row.existing_parent_task_id

        if parent_id == 0:
            continue

        if task_service.get_by_id(parent_id) is None:
            raise ValueError(
                f"Row {row.excel_row}: existing parent "
                f"task ID {parent_id} does not exist."
            )


def import_tasks(
    *,
    rows: list[ImportRow],
    source_filename: str,
    task_service: TaskService,
    activity_service: TaskActivityService,
) -> list[dict[str, object]]:
    pending = {
        row.upload_ref: row
        for row in rows
    }

    created_id_map: dict[int, int] = {}
    results: list[dict[str, object]] = []

    while pending:
        progress_made = False

        for upload_ref, row in list(
            pending.items()
        ):
            if row.parent_upload_ref > 0:
                parent_task_id = created_id_map.get(
                    row.parent_upload_ref
                )

                if parent_task_id is None:
                    continue
            else:
                parent_task_id = (
                    row.existing_parent_task_id
                )

            task = TaskCreate(
                parent_task_id=parent_task_id,
                short_description=(
                    row.short_description
                ),
                description=row.description,
                status=row.status,
                priority=row.priority,
                task_type=row.task_type,
                category=row.category,
                tags=row.tags,
                planned_start_date=(
                    row.planned_start_date
                ),
                planned_end_date=(
                    row.planned_end_date
                ),
                due_date=row.due_date,
                estimated_effort_hours=(
                    row.estimated_effort_hours
                ),
                actual_effort_hours=0,
                remaining_effort_hours=(
                    row.estimated_effort_hours
                ),
                progress_percentage=0,
                assigned_to=row.assigned_to,
                team=row.team,
                created_by=row.created_by,
            )

            created_task = task_service.create(task)

            created_id_map[upload_ref] = (
                created_task.task_id
            )

            activity_service.create(
                TaskActivityCreate(
                    task_id=created_task.task_id,
                    activity_type=(
                        ActivityType.SYSTEM
                    ),
                    title="Created by Excel Import",
                    message=(
                        f"Task created through Excel import.\n\n"
                        f"Source file: {source_filename}\n"
                        f"Upload reference: {upload_ref}\n"
                        f"Imported on: "
                        f"{datetime.now():%d %b %Y %I:%M %p}"
                    ),
                    created_by=row.created_by,
                )
            )

            results.append(
                {
                    "upload_ref": upload_ref,
                    "task_id": created_task.task_id,
                    "task_number": (
                        created_task.task_number
                    ),
                    "parent_task_id": parent_task_id,
                    "short_description": (
                        created_task.short_description
                    ),
                    "result": "Created",
                }
            )

            del pending[upload_ref]
            progress_made = True

        if not progress_made:
            unresolved = ", ".join(
                str(value)
                for value in sorted(pending)
            )

            raise ValueError(
                "Unable to resolve parent hierarchy. "
                "Possible circular parent references for "
                f"Upload Ref values: {unresolved}"
            )

    return results


def write_results(
    *,
    results: list[dict[str, object]],
    output_path: Path,
) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Import Results"

    headers = [
        "Upload Ref",
        "Created Task ID",
        "Created Task Number",
        "Parent Task ID",
        "Short Description",
        "Result",
    ]

    worksheet.append(headers)

    for result in results:
        worksheet.append(
            [
                result["upload_ref"],
                result["task_id"],
                result["task_number"],
                result["parent_task_id"],
                result["short_description"],
                result["result"],
            ]
        )

    fill = PatternFill(
        fill_type="solid",
        fgColor="1F4E78",
    )

    font = Font(
        color="FFFFFF",
        bold=True,
    )

    for cell in worksheet[1]:
        cell.fill = fill
        cell.font = font

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = (
        worksheet.dimensions
    )

    widths = [14, 18, 22, 18, 45, 14]

    for index, width in enumerate(
        widths,
        start=1,
    ):
        worksheet.column_dimensions[
            chr(64 + index)
        ].width = width

    workbook.save(output_path)


def main() -> None:
    parser = argparse.ArgumentParser(
        description=(
            "Import task hierarchy from Excel."
        )
    )

    parser.add_argument(
        "--file",
        required=True,
        help="Path to the Excel workbook.",
    )

    parser.add_argument(
        "--result-file",
        required=False,
        help=(
            "Optional path for the import-result workbook."
        ),
    )

    arguments = parser.parse_args()

    input_path = Path(arguments.file).resolve()

    if not input_path.exists():
        raise FileNotFoundError(
            f"Input file does not exist: {input_path}"
        )

    rows = read_import_rows(input_path)

    database = SessionLocal()

    try:
        task_service, activity_service = (
            create_services(database)
        )

        validate_existing_parents(
            rows,
            task_service,
        )

        print(
            f"Validation successful. "
            f"{len(rows)} task(s) will be created."
        )

        results = import_tasks(
            rows=rows,
            source_filename=input_path.name,
            task_service=task_service,
            activity_service=activity_service,
        )

        if arguments.result_file:
            result_path = Path(
                arguments.result_file
            ).resolve()
        else:
            timestamp = datetime.now().strftime(
                "%Y%m%d_%H%M%S"
            )

            result_path = input_path.with_name(
                f"import_results_{timestamp}.xlsx"
            )

        write_results(
            results=results,
            output_path=result_path,
        )

        print()
        print(
            f"{len(results)} task(s) created successfully."
        )

        for result in results:
            print(
                f"Upload Ref {result['upload_ref']} "
                f"-> Task ID {result['task_id']} "
                f"({result['task_number']})"
            )

        print()
        print(
            f"Result workbook: {result_path}"
        )

    finally:
        database.close()


if __name__ == "__main__":
    main()